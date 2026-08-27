#!/usr/bin/env python3
import json
import re
import sys
from datetime import datetime, time
from pathlib import Path
from openpyxl import load_workbook

INPUT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('at.tension-2026.xlsx')
OUTPUT = Path(sys.argv[2]) if len(sys.argv) > 2 else Path('data.json')
DAY_ORDER = ['Do', 'Fr', 'Sa', 'So']


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def as_time(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%H:%M')
    if isinstance(value, time):
        return value.strftime('%H:%M')
    m = re.match(r'^(\d{1,2}):(\d{2})', str(value).strip())
    return f'{int(m.group(1)):02d}:{m.group(2)}' if m else None


def norm_title(value):
    s = clean(value) or ''
    return (s.replace('’', "'").replace('‘', "'").replace('–', '-').replace('—', '-')
             .replace('\u00a0', ' ').strip().casefold())


wb = load_workbook(INPUT, data_only=True)
shows_ws = wb['Shows']
calendar_ws = wb['Show-Kalender']

header_row = None
headers = {}
for r in range(1, min(shows_ws.max_row, 20) + 1):
    values = [clean(shows_ws.cell(r, c).value) for c in range(1, shows_ws.max_column + 1)]
    if 'Titel' in values and 'Typ' in values:
        header_row = r
        headers = {value: i + 1 for i, value in enumerate(values) if value}
        break
if header_row is None:
    raise RuntimeError('Header row with Titel and Typ not found in Shows sheet.')

meta_by_title = {}
for r in range(header_row + 1, shows_ws.max_row + 1):
    title = clean(shows_ws.cell(r, headers['Titel']).value)
    if not title:
        continue
    meta_by_title[norm_title(title)] = {
        'type': clean(shows_ws.cell(r, headers['Typ']).value),
        'organizer': clean(shows_ws.cell(r, headers.get('Veranstalter:in / Gruppe')).value) if headers.get('Veranstalter:in / Gruppe') else None,
        'shortDescription': clean(shows_ws.cell(r, headers.get('Kurzbeschreibung')).value) if headers.get('Kurzbeschreibung') else None,
        'contentWarning': clean(shows_ws.cell(r, headers.get('ContentWarningHinweis')).value) if headers.get('ContentWarningHinweis') else None,
        'performers': clean(shows_ws.cell(r, headers.get('Von und mit')).value) if headers.get('Von und mit') else None,
        'duration': clean(shows_ws.cell(r, headers.get('Dauer')).value) if headers.get('Dauer') else None,
        'minAge': clean(shows_ws.cell(r, headers.get('Mindestalter')).value) if headers.get('Mindestalter') else None,
        'language': clean(shows_ws.cell(r, headers.get('Sprache')).value) if headers.get('Sprache') else None,
        'stage': clean(shows_ws.cell(r, headers.get('Bühne')).value) if headers.get('Bühne') else None,
    }

calendar_header_row = None
day_columns = {}
for r in range(1, min(calendar_ws.max_row, 10) + 1):
    for c in range(1, calendar_ws.max_column + 1):
        value = clean(calendar_ws.cell(r, c).value)
        if value in DAY_ORDER:
            day_columns[value] = c
    if len(day_columns) == 4:
        calendar_header_row = r
        break
if calendar_header_row is None:
    raise RuntimeError('Day columns Do, Fr, Sa, So not found in Show-Kalender sheet.')

current_time = None
entries = []
for r in range(calendar_header_row + 1, calendar_ws.max_row + 1):
    parsed_time = as_time(calendar_ws.cell(r, 1).value)
    if parsed_time:
        current_time = parsed_time
    if not current_time:
        continue

    for day in DAY_ORDER:
        title = clean(calendar_ws.cell(r, day_columns[day]).value)
        if not title:
            continue
        meta = meta_by_title.get(norm_title(title), {})
        entries.append({
            'title': title,
            'day': day,
            'time': current_time,
            'type': meta.get('type'),
            'stage': meta.get('stage'),
            'organizer': meta.get('organizer'),
            'duration': meta.get('duration'),
            'minAge': meta.get('minAge'),
            'language': meta.get('language'),
            'shortDescription': meta.get('shortDescription'),
            'contentWarning': meta.get('contentWarning'),
            'performers': meta.get('performers'),
        })

day_index = {day: i for i, day in enumerate(DAY_ORDER)}
entries.sort(key=lambda x: (day_index[x['day']], x['time'], x['title'].casefold()))
OUTPUT.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding='utf-8')
print(f'Created {OUTPUT} with {len(entries)} show appointments.')
